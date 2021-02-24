from openpyxl.chart import (
    ScatterChart,
    Reference,
    LineChart,
    Series,
    AreaChart
)
from openpyxl.utils import get_column_letter


def create_scatter(df, wb, sheet_name):
    df_columns = df.columns.tolist()
    ws = wb[sheet_name]
    chart = ScatterChart()
    chart.title = sheet_name
    chart.height = 15
    chart.width = 30

    chart.x_axis.scaling.max = max(df['timestamps'])
    chart.x_axis.scaling.min = min(df['timestamps'])
    chart.x_axis.title = df_columns[1]

    for i in df_columns:
        if i in sheet_name:
            chart.x_axis.title = i
    chart.y_axis.title = 'intensity_of_emotion'
    start_row = 2
    end_row = len(df) + 1

    x_values = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)
    for i in range(df_columns.index('happy') + 1, df_columns.index('confused') + 2):
        values = Reference(ws, min_col=i, min_row=1, max_row=end_row)
        series = Series(values, x_values, title_from_data=True)
        chart.series.append(series)

    letter = get_column_letter(len(df_columns) + 2)
    ws.add_chart(chart, f"{letter}1")


def create_line(df, wb, sheet_name, startcol=0):
    df_columns = df.columns.tolist()
    ws = wb[sheet_name]

    # Chart with date axis
    chart = LineChart()
    chart.y_axis.title = 'intensity_of_emotion'
    chart.grouping = "standard"

    if sheet_name != 'max_by_resp':
        chart.title = sheet_name
        chart.x_axis.title = df_columns[1]
        chart.height = 15
        chart.width = 30
        data = Reference(ws, min_col=startcol + 3, min_row=1, max_col=len(df_columns) + startcol,
                         max_row=len(df) + 1)
        chart.add_data(data, titles_from_data=True)
        episodes = Reference(ws, min_col=startcol + 2, min_row=2, max_row=len(df) + 1)
    else:
        chart.height = 10
        chart.width = 20
        chart.title = df.at[0, 'user_id']
        chart.x_axis.title = df_columns[2]
        data = Reference(ws, min_col=startcol + 4, min_row=1, max_col=startcol + len(df_columns),
                         max_row=len(df) + 1)
        chart.add_data(data, titles_from_data=True)
        episodes = Reference(ws, min_col=startcol + 3, min_row=2, max_row=len(df) + 1)

    chart.set_categories(episodes)
    for i in chart.series:
        i.smooth = True

    letter = get_column_letter(startcol + 1)
    ws.add_chart(chart, f"{letter}{len(df) + 3}")


def area_max(df, wb, sheet_name):
    df_columns = df.columns.tolist()
    ws = wb[sheet_name]

    chart = AreaChart()
    chart.height = 15
    chart.width = 30
    chart.x_axis.title = df_columns[1]
    chart.y_axis.title = None
    chart.grouping = "percentStacked"
    chart.title = sheet_name

    episodes = Reference(ws, min_col=df_columns.index('episode') + 1, min_row=2, max_row=len(df) + 1)
    data = Reference(ws, min_col=df_columns.index('happy') + 1, min_row=1, max_col=len(df_columns), max_row=len(df) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(episodes)

    ws.add_chart(chart, f"A{len(df) + 3}")


if __name__ == '__main__':
    # create_chart_by_timestamps()
    # create_chart_by_episodes()
    # area_max()
    pass
